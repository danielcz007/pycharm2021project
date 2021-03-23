# -*- coding: utf-8 -*-
# @Time    : 2021/1/31 19:03
# @Author  : Daniel
'''
    Yaml--json-excel
    yaml:是一个可读性高，用来表达数据“序列化”的格式，常常作为配置文件使用
    json:是一个轻量级的数据交换语言，该语言以易于阅读的文字基础，用来传输由 属性值或者序列性的值组成的数据对象
    excel:由直观的界面、出色的计算功能和图标工具，是一款电子制表软件
        excel: http://www.python-excel.org  ////  "openpyxl":  https://openpyxl.readthedocs.io/en/stable/  查看实例
'''

#openpyxl sample code:
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])


# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

'''
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title='my_sheet')
for row in range(4,8):
    for col in range(4,8):


wb.save(filename = dest_filename)
'''